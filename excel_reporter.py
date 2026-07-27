import os
import pandas as pd
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_report(analyzed_results: List[Dict], output_filepath: str):
    """
    Generates a beautifully formatted multi-tab Excel report containing:
    1. Bankipur_Consti_Opinions (strict constituency level voters)
    2. All_Respondent_Opinions (complete raw voter list)
    3. Video_Summaries (video level metadata)
    4. Aggregate_Dashboard (macro statistics)
    """
    bankipur_respondent_rows = []
    all_respondent_rows = []
    video_summary_rows = []
    
    party_counts_bankipur = {}
    issue_counts_bankipur = {}
    total_bankipur_respondents = 0

    for res in analyzed_results:
        meta = res.get("metadata", {})
        analysis = res.get("analysis")
        
        video_title = meta.get("title", "")
        channel = meta.get("channel", "")
        upload_date = meta.get("upload_date", "")
        video_url = meta.get("url", "")
        is_bankipur = meta.get("is_bankipur_constituency", True)
        audit_reason = meta.get("constituency_audit_reason", "")

        if not analysis:
            continue

        if isinstance(analysis, dict):
            resp_list = analysis.get("respondents", [])
            overall_summary = analysis.get("overall_video_summary", "")
        else:
            resp_list = getattr(analysis, "respondents", [])
            overall_summary = getattr(analysis, "overall_video_summary", "")

        v_party_counts = {}

        for resp in resp_list:
            if isinstance(resp, dict):
                p_party = resp.get("preferred_party", "Others")
                p_certainty = resp.get("stance_certainty", "Undecided")
                p_reason = resp.get("key_reason", "")
                p_issues = resp.get("key_issues", [])
                p_quote_orig = resp.get("quote_original", "")
                p_quote_eng = resp.get("quote_english", "")
                p_demo = resp.get("demographics_or_context", "")
                p_id = resp.get("respondent_id", "Respondent")
            else:
                p_party = getattr(resp, "preferred_party", "Others")
                p_certainty = getattr(resp, "stance_certainty", "Undecided")
                p_reason = getattr(resp, "key_reason", "")
                p_issues = getattr(resp, "key_issues", [])
                p_quote_orig = getattr(resp, "quote_original", "")
                p_quote_eng = getattr(resp, "quote_english", "")
                p_demo = getattr(resp, "demographics_or_context", "")
                p_id = getattr(resp, "respondent_id", "Respondent")

            row_data = {
                "Bankipur Specific?": "YES (Bankipur)" if is_bankipur else "NO (General Bihar/Lok Sabha)",
                "Video Title": video_title,
                "Channel": channel,
                "Upload Date": upload_date,
                "Respondent ID": p_id,
                "Preferred Party": p_party,
                "Stance Certainty": p_certainty,
                "Key Reason": p_reason,
                "Key Issues": ", ".join(p_issues) if isinstance(p_issues, list) else str(p_issues),
                "Quote (Original)": p_quote_orig,
                "Quote (English)": p_quote_eng,
                "Demographics / Context": p_demo,
                "Video URL": video_url
            }

            all_respondent_rows.append(row_data)

            if is_bankipur:
                total_bankipur_respondents += 1
                bankipur_respondent_rows.append(row_data)
                party_counts_bankipur[p_party] = party_counts_bankipur.get(p_party, 0) + 1

                if isinstance(p_issues, list):
                    for issue in p_issues:
                        clean_issue = str(issue).strip().title()
                        if clean_issue:
                            issue_counts_bankipur[clean_issue] = issue_counts_bankipur.get(clean_issue, 0) + 1

            v_party_counts[p_party] = v_party_counts.get(p_party, 0) + 1

        breakdown_str = ", ".join([f"{k}: {v}" for k, v in v_party_counts.items()])
        video_summary_rows.append({
            "Bankipur Specific?": "YES" if is_bankipur else "NO",
            "Audit Reason": audit_reason,
            "Video Title": video_title,
            "Channel": channel,
            "Upload Date": upload_date,
            "Total Respondents": len(resp_list),
            "Party Preference Breakdown": breakdown_str,
            "Overall Video Summary": overall_summary,
            "Video URL": video_url
        })

    df_bankipur = pd.DataFrame(bankipur_respondent_rows)
    df_all = pd.DataFrame(all_respondent_rows)
    df_video_summaries = pd.DataFrame(video_summary_rows)

    # Prepare Bankipur Aggregate Share
    party_df_data = []
    for party, count in sorted(party_counts_bankipur.items(), key=lambda x: x[1], reverse=True):
        pct = (count / total_bankipur_respondents * 100) if total_bankipur_respondents > 0 else 0
        party_df_data.append({
            "Political Party": party,
            "Respondent Count": count,
            "Percentage Share (%)": round(pct, 2)
        })

    issues_df_data = []
    for issue, count in sorted(issue_counts_bankipur.items(), key=lambda x: x[1], reverse=True):
        issues_df_data.append({
            "Key Issue": issue,
            "Times Mentioned": count
        })

    with pd.ExcelWriter(output_filepath, engine="openpyxl") as writer:
        if not df_bankipur.empty:
            df_bankipur.to_excel(writer, sheet_name="Bankipur_Consti_Opinions", index=False)
        if not df_all.empty:
            df_all.to_excel(writer, sheet_name="All_Respondent_Opinions", index=False)
        if not df_video_summaries.empty:
            df_video_summaries.to_excel(writer, sheet_name="Video_Summaries", index=False)

        workbook = writer.book
        dashboard_sheet = workbook.create_sheet(title="Aggregate_Dashboard")
        
        # Title Banner
        dashboard_sheet.merge_cells("A1:E1")
        title_cell = dashboard_sheet["A1"]
        title_cell.value = "BANKIPUR CONSTITUENCY ELECTION OPINION DASHBOARD"
        title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        dashboard_sheet.row_dimensions[1].height = 35

        dashboard_sheet["A3"] = "Bankipur Specific Videos Analyzed:"
        dashboard_sheet["B3"] = len([r for r in video_summary_rows if r["Bankipur Specific?"] == "YES"])
        dashboard_sheet["A4"] = "Bankipur Respondents Interviewed:"
        dashboard_sheet["B4"] = total_bankipur_respondents
        dashboard_sheet["A5"] = "Filtered Out General Bihar Videos:"
        dashboard_sheet["B5"] = len([r for r in video_summary_rows if r["Bankipur Specific?"] == "NO"])

        for cell_ref in ["A3", "A4", "A5"]:
            dashboard_sheet[cell_ref].font = Font(bold=True)

        dashboard_sheet["A7"] = "Bankipur Constituency Party Preference Share"
        dashboard_sheet["A7"].font = Font(size=13, bold=True, color="1F4E78")
        
        headers_party = ["Political Party", "Respondent Count", "Share (%)"]
        for col_idx, h in enumerate(headers_party, start=1):
            cell = dashboard_sheet.cell(row=8, column=col_idx)
            cell.value = h
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        curr_row = 9
        for p_row in party_df_data:
            dashboard_sheet.cell(row=curr_row, column=1, value=p_row["Political Party"])
            dashboard_sheet.cell(row=curr_row, column=2, value=p_row["Respondent Count"])
            dashboard_sheet.cell(row=curr_row, column=3, value=f"{p_row['Percentage Share (%)']}%")
            curr_row += 1

        curr_row += 2
        dashboard_sheet.cell(row=curr_row, column=1, value="Top Bankipur Constituency Issues").font = Font(size=13, bold=True, color="1F4E78")
        curr_row += 1
        
        headers_issues = ["Key Issue", "Times Mentioned"]
        for col_idx, h in enumerate(headers_issues, start=1):
            cell = dashboard_sheet.cell(row=curr_row, column=col_idx)
            cell.value = h
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        curr_row += 1
        for i_row in issues_df_data:
            dashboard_sheet.cell(row=curr_row, column=1, value=i_row["Key Issue"])
            dashboard_sheet.cell(row=curr_row, column=2, value=i_row["Times Mentioned"])
            curr_row += 1

    wb = openpyxl.load_workbook(output_filepath)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name in ["Bankipur_Consti_Opinions", "All_Respondent_Opinions", "Video_Summaries"]:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 28

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len and not type(cell).__name__ == 'MergedCell':
                    max_len = min(len(val_str), 50)
                if sheet_name in ["Bankipur_Consti_Opinions", "All_Respondent_Opinions", "Video_Summaries"] and cell.row > 1:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(output_filepath)
    return output_filepath
